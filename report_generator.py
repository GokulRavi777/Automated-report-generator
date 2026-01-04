import pandas as pd
from datetime import datetime, timedelta
import os
import json
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment

def load_config():
    try:
        with open('config.json', 'r') as f:
            return json.load(f)
    except FileNotFoundError:
        return {
            "input_file": "customers.csv",
            "output_dir": "reports",
            "date_column": "date",
            "reports": [
                {"name": "daily", "days": 1},
                {"name": "weekly", "days": 7},
                {"name": "monthly", "days": 30}
            ]
        }

def style_worksheet(ws):
    # Dim header font and fill
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_fill = PatternFill("solid", fgColor="4F81BD")
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                         top=Side(style='thin'), bottom=Side(style='thin'))
    
    # Style headers
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Auto-adjust column widths
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter # Get the column name
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width
        
    # Add borders to all data
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border

def create_charts(wb, df):
    if 'service_type' not in df.columns or 'amount' not in df.columns:
        return

    # Create a 'Dashboard' sheet
    ws_dash = wb.create_sheet("Dashboard", 0)
    ws_dash.sheet_view.showGridLines = False
    
    # --- PREPARE DATA FOR CHARTS ---
    # Pivot for Amount by Service (Bar Chart)
    pivot_amount = df.pivot_table(index='service_type', values='amount', aggfunc='sum').reset_index()
    pivot_amount = pivot_amount.sort_values(by='amount', ascending=False)
    
    # Pivot for Count by Service (Pie Chart)
    pivot_count = df['service_type'].value_counts().reset_index()
    pivot_count.columns = ['service_type', 'count']

    # Write data to a hidden 'Data_Stats' sheet
    ws_stats = wb.create_sheet("Data_Stats")
    ws_stats.sheet_state = 'hidden'
    
    # Write Amount Data
    ws_stats.append(["Service Type", "Revenue"])
    for r in dataframe_to_rows(pivot_amount, index=False, header=False):
        ws_stats.append(r)
    amount_len = len(pivot_amount)

    # Write Count Data (offset by a few columns)
    ws_stats.cell(row=1, column=4, value="Service Type")
    ws_stats.cell(row=1, column=5, value="Project Count")
    for idx, r in enumerate(dataframe_to_rows(pivot_count, index=False, header=False)):
        ws_stats.cell(row=idx+2, column=4, value=r[0])
        ws_stats.cell(row=idx+2, column=5, value=r[1])
    count_len = len(pivot_count)

    # --- CREATE BAR CHART (Revenue) ---
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 10
    bar_chart.title = "Revenue by Service Type"
    bar_chart.y_axis.title = 'Amount ($)'
    bar_chart.x_axis.title = 'Service'
    bar_chart.height = 10
    bar_chart.width = 18

    data = Reference(ws_stats, min_col=2, min_row=1, max_row=amount_len+1, max_col=2)
    cats = Reference(ws_stats, min_col=1, min_row=2, max_row=amount_len+1)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    
    ws_dash.add_chart(bar_chart, "B2")

    # --- CREATE PIE CHART (Project Count) ---
    pie_chart = PieChart()
    pie_chart.title = "Project Distribution"
    pie_chart.height = 10
    pie_chart.width = 10
    
    data_pie = Reference(ws_stats, min_col=5, min_row=1, max_row=count_len+1)
    cats_pie = Reference(ws_stats, min_col=4, min_row=2, max_row=count_len+1)
    pie_chart.add_data(data_pie, titles_from_data=True)
    pie_chart.set_categories(cats_pie)
    
    ws_dash.add_chart(pie_chart, "L2")

def main():
    config = load_config()
    input_file = config['input_file']
    output_dir = config['output_dir']
    date_col = config['date_column']
    
    os.makedirs(output_dir, exist_ok=True)
    
    print(f"Loading data from {input_file}...")
    df = pd.read_csv(input_file)
    df[date_col] = pd.to_datetime(df[date_col])
    
    now = datetime.now()
    
    for report in config['reports']:
        name = report['name']
        days = report['days']
        
        start_date = now - timedelta(days=days)
        filtered_df = df[df[date_col] >= start_date]
        
        if filtered_df.empty:
            print(f"⚠️ No data found for {name} report (Last {days} days).")
            continue
            
        output_path = os.path.join(output_dir, f"{name}_report.xlsx")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Raw Data"
        
        # Write data
        for r in dataframe_to_rows(filtered_df, index=False, header=True):
            ws.append(r)
            
        # Apply Styling
        style_worksheet(ws)
        
        # Create Dashboard with Charts
        create_charts(wb, filtered_df)
        
        wb.save(output_path)
        print(f" Generated {name} report -> {output_path} ({len(filtered_df)} rows)")
        
    print("\n Report generation complete!")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"Critical Error: {e}")
        import sys
        sys.exit(1)
